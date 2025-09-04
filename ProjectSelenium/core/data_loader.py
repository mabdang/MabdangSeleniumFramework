import os
import openpyxl
from core.yaml_reader import YAMLReader
from core.csv_reader import CSVReader

class DataLoader:
    @staticmethod
    def load_testcases(folder):
        """
        Mencari testcases.yaml / testcases.csv / testcases.xlsx di folder,
        prioritas YAML > CSV > XLSX.
        """
        paths = [
            os.path.join(folder, "testcases.yaml"),
            os.path.join(folder, "testcases.yml"),
            os.path.join(folder, "testcases.csv"),
            os.path.join(folder, "testcases.xlsx"),
        ]

        for path in paths:
            if os.path.exists(path):
                ext = os.path.splitext(path)[1].lower()
                if ext in [".yaml", ".yml"]:
                    print(f"[INFO] Load testcases YAML: {path}")
                    return YAMLReader.read(path)
                elif ext == ".csv":
                    print(f"[INFO] Load testcases CSV: {path}")
                    return CSVReader.read_testcases(path)
                elif ext == ".xlsx":
                    print(f"[INFO] Load testcases XLSX: {path}")
                    wb = openpyxl.load_workbook(path)
                    sheet = wb.active
                    keys = [cell.value for cell in sheet[1]]

                    testcases = {}
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(keys, row))
                        case_id = row_dict["CaseID"]

                        step = {
                            "StepID": row_dict["StepID"],
                            "Action": row_dict["Action"],
                            "Locator": row_dict.get("Locator"),
                            "TestData": row_dict.get("TestData"),
                            "Expected": row_dict.get("Expected"),
                        }

                        if case_id not in testcases:
                            testcases[case_id] = {
                                "CaseID": case_id,
                                "Title": row_dict["Title"],
                                "ScenarioType": row_dict["ScenarioType"],
                                "Run": str(row_dict["Run"]).lower() in ("true", "yes", "1"),
                                "TestSteps": [],
                            }

                        testcases[case_id]["TestSteps"].append(step)

                    # bungkus biar konsisten
                    return {"test_cases": list(testcases.values())}
     

        raise FileNotFoundError(f"Tidak ditemukan testcases di {folder}")

    @staticmethod
    def load_locators(folder):
        """
        Mencari locators.yaml / locators.csv / locators.xlsx di folder,
        prioritas YAML > CSV > XLSX.
        """
        paths = [
            os.path.join(folder, "locators.yaml"),
            os.path.join(folder, "locators.yml"),
            os.path.join(folder, "locators.csv"),
            os.path.join(folder, "locators.xlsx"),
        ]

        for path in paths:
            if os.path.exists(path):
                ext = os.path.splitext(path)[1].lower()
                if ext in [".yaml", ".yml"]:
                    print(f"[INFO] Load locators YAML: {path}")
                    data = YAMLReader.read(path)
                    return data.get("locators", data)
                elif ext == ".csv":
                    print(f"[INFO] Load locators CSV: {path}")
                    return CSVReader.read_locators(path)
                elif ext == ".xlsx":
                    print(f"[INFO] Load locators XLSX: {path}")
                    wb = openpyxl.load_workbook(path)
                    sheet = wb.active
                    keys = [cell.value for cell in sheet[1]]
                    locators = {}
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(keys, row))
                        name = row_dict["LocatorName"]
                        locators[name] = {
                            "LocatorType": row_dict.get("LocatorType"),
                            "LocatorValue": row_dict.get("LocatorValue"),
                            "Description": row_dict.get("Description", ""),
                        }
                    return locators

        raise FileNotFoundError(f"Tidak ditemukan locators di {folder}")
